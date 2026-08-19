"""Encrypt 2026 H1 HK earnings analysis for static hosting.

Usage:
  python3 encrypt_hk_h1_2026.py <unified_password> [xlsx_path]
"""
from __future__ import annotations

import base64
import json
import os
import sys
from datetime import date, datetime

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage: python3 encrypt_hk_h1_2026.py <unified_password> [xlsx]")
    sys.exit(1)

PASSWORD = sys.argv[1]
XLSX = sys.argv[2] if len(sys.argv) > 2 else os.path.expanduser(
    "~/.hermes/cache/documents/doc_1ce62c8e21fb_2026H1_HK.xlsx"
)
BASE = os.path.dirname(os.path.abspath(__file__))
DST = os.path.join(BASE, "hk_h1_2026_data.enc")

KEYS = [
    "code", "name", "period", "date", "type", "category", "rank", "priority",
    "score", "rationale", "thesis", "revenue", "rev_yoy", "gm", "gm_chg",
    "op_profit", "op_yoy", "ni", "ni_yoy", "ocf", "leverage", "dividend",
    "margin_trend", "leading", "catalyst", "risk", "outlook", "cash_quality",
]


def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float):
        return round(v, 4) if v == v else None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def encrypt_bytes(plaintext: bytes, password: str) -> str:
    salt = os.urandom(16)
    nonce = os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100_000)
    key = kdf.derive(password.encode())
    ct = AESGCM(key).encrypt(nonce, plaintext, None)
    return base64.b64encode(salt + nonce + ct).decode()


wb = load_workbook(XLSX, data_only=True)
ws = wb.active
rows = []
for raw in ws.iter_rows(min_row=2, values_only=True):
    rec = {k: clean(raw[i] if i < len(raw) else None) for i, k in enumerate(KEYS)}
    if not rec.get("code"):
        continue
    rec["code"] = str(rec["code"]).zfill(5)
    rows.append(rec)

payload = json.dumps(
    {"generated": "2026-08-19", "source": "2026H1_HK.xlsx", "count": len(rows), "rows": rows},
    ensure_ascii=False,
    separators=(",", ":"),
).encode("utf-8")
enc = encrypt_bytes(payload, PASSWORD)
with open(DST, "w", encoding="utf-8") as f:
    f.write(enc)
print(f"rows={len(rows)} bytes={len(payload)} enc={os.path.getsize(DST)} -> {DST}")
